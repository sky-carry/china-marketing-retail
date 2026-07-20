# -*- coding: utf-8 -*-
"""Excel 导出：从核对视图现查现生成工作簿，按数据版本（ETag）缓存。

两个独立导出，各对应看板的一个 Tab：
  kind='recon' 客户/门店核对：客户汇总 + 门店汇总 + 核对明细 + 未匹配门店 + 未匹配商品
  kind='guard' 网点保障：网点保障汇总 + 网点保障明细
"""
import io
import threading
from typing import Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..db import get_conn
from . import recon


def _maintain_rate(platform, bojun):
    """维护率：平台 ≥ 伯俊 → 100；否则 平台/伯俊×100；无可比基数 → None"""
    if platform is None or bojun is None:
        return None
    platform, bojun = float(platform), float(bojun)
    if platform >= bojun:
        return 100.0
    if bojun > 0:
        return round(100.0 * platform / bojun, 1)
    return None


# 客户/门店汇总的聚合列，口径与看板树表完全一致：
#   差异 = Σ(平台−伯俊)，仅累计伯俊有记录、且非「平台虚拟库存」的行
#   维护率 = LEAST(100, Σmin(平台,伯俊) ÷ Σ伯俊)，仅伯俊>0 且平台已上翻的行
_AGG_COLS = """
  COALESCE(sum(jd_diff) FILTER (WHERE bojun_qty IS NOT NULL AND flag <> '平台虚拟库存'), 0) AS jd_diff,
  LEAST(100, round(100.0 * sum(LEAST(jd_qty, bojun_qty)) FILTER (WHERE bojun_qty > 0 AND jd_qty IS NOT NULL)
        / NULLIF(sum(bojun_qty) FILTER (WHERE bojun_qty > 0 AND jd_qty IS NOT NULL), 0), 1)) AS jd_rate,
  COALESCE(sum(mt_diff) FILTER (WHERE bojun_qty IS NOT NULL AND flag <> '平台虚拟库存'), 0) AS mt_diff,
  LEAST(100, round(100.0 * sum(LEAST(mt_qty, bojun_qty)) FILTER (WHERE bojun_qty > 0 AND mt_qty IS NOT NULL)
        / NULLIF(sum(bojun_qty) FILTER (WHERE bojun_qty > 0 AND mt_qty IS NOT NULL), 0), 1)) AS mt_rate"""


# 门店 → 京东/美团门店ID（v_dim_store 每店可能多行，先按店名去重取一个，避免 join 放大合计）
_STORE_IDS = """(SELECT store_name, max(jd_id) AS jd_id, max(meituan_id) AS meituan_id
                 FROM v_dim_store WHERE is_active GROUP BY store_name)"""


def _build_recon(sheet):
    """客户/门店核对相关 sheet（与「客户/门店核对」Tab 一致）。"""
    sheet('客户汇总',
          ['客户名称', '门店数', '伯俊库存', '京东库存', '京东差异', '京东维护率%',
           '美团库存', '美团差异', '美团维护率%'],
          f"""SELECT customer_name, store_cnt, bojun_sum, jd_sum, jd_diff, jd_rate,
                     mt_sum, mt_diff, mt_rate FROM (
                SELECT customer_name, count(DISTINCT store_name) AS store_cnt,
                       COALESCE(sum(bojun_qty), 0) AS bojun_sum,
                       COALESCE(sum(jd_qty), 0)    AS jd_sum,
                       COALESCE(sum(mt_qty), 0)    AS mt_sum,{_AGG_COLS}
                FROM v_recon_detail GROUP BY customer_name
              ) t ORDER BY abs(jd_diff) + abs(mt_diff) DESC""",
          widths=[32, 8, 10, 10, 10, 12, 10, 10, 12])
    sheet('门店汇总',
          ['客户名称', '门店名称', '伯俊库存', '京东门店ID', '京东库存', '京东差异',
           '京东维护率%', '美团门店ID', '美团库存', '美团差异', '美团维护率%'],
          f"""SELECT customer_name, store_name, bojun_sum, jd_id, jd_sum, jd_diff, jd_rate,
                     meituan_id, mt_sum, mt_diff, mt_rate FROM (
                SELECT d.customer_name, d.store_name,
                       COALESCE(sum(d.bojun_qty), 0) AS bojun_sum,
                       COALESCE(sum(d.jd_qty), 0)    AS jd_sum,
                       COALESCE(sum(d.mt_qty), 0)    AS mt_sum,
                       max(ids.jd_id) AS jd_id, max(ids.meituan_id) AS meituan_id,{_AGG_COLS}
                FROM v_recon_detail d
                LEFT JOIN {_STORE_IDS} ids ON ids.store_name = d.store_name
                GROUP BY d.customer_name, d.store_name
              ) t ORDER BY customer_name, abs(jd_diff) + abs(mt_diff) DESC""",
          widths=[32, 30, 10, 12, 10, 10, 12, 12, 10, 10, 12])
    sheet('核对明细',
          ['客户名称', '门店名称', '货号', '品名', '伯俊库存', '京东门店ID', '京东库存',
           '京东差异', '京东维护率%', '美团门店ID', '美团库存', '美团差异', '美团维护率%', '状态'],
          f"""SELECT d.customer_name, d.store_name, d.product_code, d.product_name,
                     d.bojun_qty, ids.jd_id, d.jd_qty, d.jd_diff, d.mt_diff,
                     ids.meituan_id, d.mt_qty, d.flag
             FROM v_recon_detail d
             LEFT JOIN {_STORE_IDS} ids ON ids.store_name = d.store_name
             ORDER BY d.customer_name, d.store_name, d.product_code""",
          transform=lambda rows: ([r[0], r[1], r[2], r[3], r[4], r[5], r[6],
                                   r[7], _maintain_rate(r[6], r[4]),
                                   r[9], r[10], r[8], _maintain_rate(r[10], r[4]), r[11]]
                                  for r in rows),
          widths=[32, 30, 16, 30, 10, 12, 10, 10, 12, 12, 10, 10, 12, 20])
    sheet('未匹配门店',
          ['门店名称', '伯俊店仓名(待修正)', '省份', '城市', '京东ID', '美团ID'],
          """SELECT store_name, bojun_warehouse, province, city, jd_id, meituan_id
             FROM v_store_unmatched ORDER BY province, city, store_name""",
          widths=[30, 30, 12, 12, 12, 12])
    sheet('未匹配商品',
          ['平台', '平台编码', '商品名称', '条码', '涉及门店', '库存合计'],
          """SELECT platform, platform_code, product_name, barcode, store_cnt, total_qty
             FROM v_product_unmatched ORDER BY platform, store_cnt DESC""",
          widths=[8, 16, 30, 16, 10, 12])


def _build_guard(sheet):
    """网点保障相关 sheet（与「网点保障」Tab 一致：公司→网点→货号）。"""
    sheet('网点汇总',
          ['客户名称', '平台', '网点名称', '网点编码', '产品SKU', '达标产品SKU',
           '不达标产品SKU', '达标率%'],
          """SELECT customer_name, platform, outlet_name, outlet_code,
                    count(*)                          AS product_cnt,
                    count(*) FILTER (WHERE is_ok)     AS ok_cnt,
                    count(*) FILTER (WHERE NOT is_ok) AS bad_cnt,
                    round(100.0 * count(*) FILTER (WHERE is_ok) / count(*), 1) AS ok_rate
             FROM v_outlet_guard
             GROUP BY customer_name, platform, outlet_code, outlet_name
             ORDER BY ok_rate, customer_name, platform, outlet_name""",
          widths=[32, 8, 30, 14, 10, 13, 15, 10])
    sheet('货号明细',
          ['客户名称', '平台', '网点名称', '网点编码', '货号', '品名',
           '线下伯俊总和', '线上本网点库存', '缺口', '状态'],
          """SELECT customer_name, platform, outlet_name, outlet_code,
                    product_code, product_name, offline_qty, outlet_qty, gap,
                    CASE WHEN is_ok THEN '达标' ELSE '不足' END
             FROM v_outlet_guard
             ORDER BY customer_name, platform, outlet_name, is_ok, gap, product_code""",
          widths=[32, 8, 30, 14, 16, 30, 12, 15, 10, 10])


_BUILDERS = {'recon': _build_recon, 'guard': _build_guard}


def build_xlsx(kind: str = 'recon') -> bytes:
    wb = Workbook(write_only=True)
    with get_conn() as conn:
        cur = conn.cursor()

        def sheet(name, headers, sql, transform=None, widths=None):
            cur.execute(sql)
            rows = cur.fetchall()
            ws = wb.create_sheet(name)
            if widths:                       # 按内容预设列宽，下载后无需手动拉宽
                for i, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w
            ws.append(headers)
            for r in (transform(rows) if transform else rows):
                ws.append(list(r))

        _BUILDERS.get(kind, _build_recon)(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_lock = threading.Lock()
_cache = {}   # kind -> {'etag', 'bytes'}


def get_xlsx(kind: str = 'recon') -> Tuple[str, bytes]:
    """跟随 /api/data 的数据版本缓存：数据变了才重新生成。每个 kind 独立缓存。"""
    if kind not in _BUILDERS:
        kind = 'recon'
    etag, _ = recon.get_data()
    etag = f'{kind}-{etag}'
    with _lock:
        c = _cache.get(kind)
        if c and c['etag'] == etag and c['bytes']:
            return etag, c['bytes']
    data = build_xlsx(kind)
    with _lock:
        _cache[kind] = {'etag': etag, 'bytes': data}
    return etag, data


# ---- 后台预热：启动时 + 每次数据变更后先把两个 Excel 建好，让下载点击命中缓存 ----
# 生成慢（recon ~3.5s / guard ~7s），不预热则每次改数据后第一次点击都要等。
_warm_lock = threading.Lock()
_warm_state = {'running': False, 'again': False}


def warm(kinds=('recon', 'guard')) -> None:
    """后台预生成缓存；合并突发调用（配置连改多行只跑最后一次）。"""
    with _warm_lock:
        if _warm_state['running']:
            _warm_state['again'] = True       # 已在跑：标记跑完再来一轮
            return
        _warm_state['running'] = True

    def _run():
        while True:
            for k in kinds:
                try:
                    get_xlsx(k)
                except Exception:              # noqa: BLE001 —— 预热失败点击时兜底重建
                    pass
            with _warm_lock:
                if not _warm_state['again']:
                    _warm_state['running'] = False
                    return
                _warm_state['again'] = False    # 期间又有变更：再跑一轮

    threading.Thread(target=_run, daemon=True).start()


recon.on_invalidate(warm)     # 数据入库 / 配置改动后自动预热
