# -*- coding: utf-8 -*-
"""ETL 后台运行器：接收看板上传的 Excel，保存到 excel/ 目录后异步执行入库。

同一时间只允许一个 ETL 任务；状态与日志保存在内存里供前端轮询。
"""
import os
import re
import sys
import time
import shutil
import threading
import subprocess
from typing import Optional

from ..config import BASE_DIR
from ..db import get_conn
from . import recon

EXCEL_DIR = os.path.join(BASE_DIR, 'excel')

# kind -> (落盘文件名, load_excel --only 的数据源键, 展示名)
UPLOAD_KINDS = {
    'jd_inventory': ('京东门店库存-新.xlsx', 'jd_inventory', '京东门店库存'),
    'meituan_inventory': ('美团门店库存.xlsx', 'meituan_inventory', '美团门店库存'),
    'bojun': ('伯俊线下库存.xlsx', 'bojun', '伯俊线下库存（临时，API 接入后弃用）'),
}

# kind -> (入库表名, 模板 sheet 名)。表头取自 load_excel.RENAME，与列校验同源
_TEMPLATE_META = {
    'jd_inventory': ('jd_store_inventory', 'Sheet1'),
    'meituan_inventory': ('meituan_store_inventory', 'Sheet1'),
    'bojun': ('bojun_offline_inventory', 'download'),   # 伯俊 loader 只认 download 这个 sheet 名
}


def _load_etl_module():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        'etl_load_excel', os.path.join(BASE_DIR, 'etl', 'load_excel.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)          # 顶层只有常量与函数定义，无副作用
    return mod


def build_template(kind: str) -> bytes:
    """生成上传模板：表头 = 入库列校验要求的中文列（与 RENAME 同源），首行加粗冻结。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    table, sheet_name = _TEMPLATE_META[kind]
    headers = [zh for zh in _load_etl_module().RENAME[table] if zh != '_blank']
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')
    ws.freeze_panes = 'A2'
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            max(12, len(str(h)) * 2 + 4)
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

_lock = threading.Lock()
_state = {
    'state': 'idle',          # idle | running | success | error
    'kind': '', 'label': '', 'filename': '', 'size': 0, 'operator': '',
    'started_at': 0.0, 'finished_at': 0.0,
    'log': '',
}


# ---------- 上传历史 ----------

def _ensure_log_table(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS upload_log (
        id bigserial PRIMARY KEY,
        created_at timestamptz NOT NULL DEFAULT now(),
        label text NOT NULL,
        filename text,
        size_bytes bigint,
        state text NOT NULL,
        message text
    )""")
    cur.execute('ALTER TABLE upload_log ADD COLUMN IF NOT EXISTS operator text')
    cur.execute("""
        COMMENT ON TABLE upload_log IS '数据上传/入库历史记录';
        COMMENT ON COLUMN upload_log.id IS '自增主键';
        COMMENT ON COLUMN upload_log.created_at IS '记录时间';
        COMMENT ON COLUMN upload_log.label IS '数据源名称（如 伯俊线下库存）';
        COMMENT ON COLUMN upload_log.filename IS '上传的原始文件名';
        COMMENT ON COLUMN upload_log.size_bytes IS '文件大小（字节）';
        COMMENT ON COLUMN upload_log.state IS '结果：success=成功 / error=失败';
        COMMENT ON COLUMN upload_log.message IS '结果信息/错误摘要（截断 500 字）';
        COMMENT ON COLUMN upload_log.operator IS '操作人（密码用户名或飞书姓名）';
    """)


def add_log(label: str, filename: str, size: int, state: str, message: str,
            operator: str = '') -> None:
    try:
        with get_conn() as conn:
            cur = conn.cursor()
            _ensure_log_table(cur)
            cur.execute(
                'INSERT INTO upload_log (label, filename, size_bytes, state, message, operator) '
                'VALUES (%s, %s, %s, %s, %s, %s)',
                (label, filename, size, state, (message or '')[:500], operator or None))
            conn.commit()
    except Exception:            # noqa: BLE001 —— 历史记录失败不影响主流程
        pass


def history(limit: int = 20) -> list:
    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_log_table(cur)
        conn.commit()
        cur.execute("""SELECT to_char(created_at, 'YYYY-MM-DD HH24:MI'), label, filename,
                              size_bytes, state, message, operator
                       FROM upload_log ORDER BY id DESC LIMIT %s""", (limit,))
        return [dict(zip(('time', 'label', 'filename', 'size', 'state', 'message', 'operator'), r))
                for r in cur.fetchall()]


def status() -> dict:
    with _lock:
        s = dict(_state)
    s['running_seconds'] = round(time.time() - s['started_at'], 1) \
        if s['state'] == 'running' else None
    return s


def save_upload(kind: str, content: bytes) -> str:
    """把上传内容写到 excel/ 目录的规范文件名（忽略用户原始文件名，杜绝路径注入）。"""
    fname = UPLOAD_KINDS[kind][0]
    os.makedirs(EXCEL_DIR, exist_ok=True)
    path = os.path.join(EXCEL_DIR, fname)
    tmp = path + '.uploading'
    with open(tmp, 'wb') as f:
        f.write(content)
    if os.path.exists(path):        # 留旧版备份，入库失败时回滚
        shutil.copy2(path, path + '.bak')
    # 原子替换，避免 ETL 读到半个文件。Windows 上目标可能被杀毒/索引/Excel
    # 短暂占用（WinError 5/32），带退避重试
    last_err = None
    for i in range(10):
        try:
            os.replace(tmp, path)
            return path
        except PermissionError as e:
            last_err = e
            time.sleep(0.5 * (i + 1))
    try:
        os.remove(tmp)
    except OSError:
        pass
    raise RuntimeError(
        f'目标文件被占用，无法写入 {fname}（可能被 Excel/WPS 打开，请关闭后重试）'
    ) from last_err


def start_etl(kind: str, filename: str = '', size: int = 0, operator: str = '') -> bool:
    """启动后台 ETL；已有任务在跑时返回 False。"""
    with _lock:
        if _state['state'] == 'running':
            return False
        _state.update(state='running', kind=kind, label=UPLOAD_KINDS[kind][2],
                      filename=filename, size=size, operator=operator,
                      started_at=time.time(), finished_at=0.0, log='')
    threading.Thread(target=_run, args=(kind,), daemon=True).start()
    return True


def _run(kind: str):
    only = UPLOAD_KINDS[kind][1]
    env = dict(os.environ, PYTHONIOENCODING='utf-8')
    try:
        proc = subprocess.Popen(
            [sys.executable, os.path.join(BASE_DIR, 'etl', 'load_excel.py'), '--only', only],
            cwd=BASE_DIR, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding='utf-8', errors='replace')
        lines = []
        for line in proc.stdout:
            lines.append(line.rstrip())
            with _lock:
                _state['log'] = '\n'.join(lines[-15:])
        code = proc.wait()
        ok = code == 0
    except Exception as e:      # noqa: BLE001 —— 后台线程必须兜住一切异常
        ok, lines = False, [f'启动 ETL 失败: {e}']
        with _lock:
            _state['log'] = '\n'.join(lines)
    if not ok:                  # 入库失败：把 excel/ 里的文件回滚成上一版，避免留下坏文件
        path = os.path.join(EXCEL_DIR, UPLOAD_KINDS[kind][0])
        bak = path + '.bak'
        if os.path.exists(bak):
            try:
                os.replace(bak, path)
                lines.append('已回滚 excel 目录中的源文件为上一版')
            except OSError as e:
                lines.append(f'源文件回滚失败: {e}')
            with _lock:
                _state['log'] = '\n'.join(lines[-15:])
    with _lock:
        _state.update(state='success' if ok else 'error', finished_at=time.time())
        label, fname, size, operator = (_state['label'], _state['filename'],
                                        _state['size'], _state['operator'])
    # 写入上传历史：成功记入库行数，失败记错误原因
    text = '\n'.join(lines)
    if ok:
        m = re.search(r'(\d+) rows', text)
        msg = f'入库 {m.group(1)} 行' if m else '入库完成'
    else:
        err_lines = [l for l in lines if l.strip() and 'Warning' not in l]
        msg = '失败: ' + (err_lines[-1] if err_lines else '未知错误')
    add_log(label.split('（')[0], fname, size, 'success' if ok else 'error', msg)
    if ok:
        recon.invalidate()      # 让 /api/data 立即出新数据
