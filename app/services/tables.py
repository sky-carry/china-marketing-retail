# -*- coding: utf-8 -*-
"""通用配置表管理（门店映射 / 京东网点 / 美团网点）：增删改查 + Excel 导入导出。

三张表都无主键，幂等补自增 id；导入=清空重灌（不删表，视图不受影响）；
任何变更后刷新 data_meta（看板数据快照时间）并失效看板缓存。
路由见 main.py 的 /api/table/{key}。
"""
import io

from ..db import get_conn
from . import recon


class TableManager:
    def __init__(self, key, table, title, subtitle, columns, display, required, key_fields):
        self.key = key
        self.table = table
        self.title = title
        self.subtitle = subtitle
        self.columns = columns            # [(en, zh, type)] 全列，导出/导入用
        self.display = display            # [(en, wide)] 界面显示/编辑的列子集
        self.required = required          # 不能为空的字段
        self.key_fields = key_fields      # 判定空行用
        self.en_cols = [c for c, _, _ in columns]
        self.zh2en = {zh: en for en, zh, _ in columns}
        self.en2zh = {en: zh for en, zh, _ in columns}
        self.types = {en: t for en, _, t in columns}
        self.editable = {en for en, _ in display}
        self.not_empty = ' OR '.join(f'{c} IS NOT NULL' for c in key_fields)

    def schema(self) -> dict:
        return {'key': self.key, 'title': self.title, 'subtitle': self.subtitle,
                'nameField': self.required, 'nameLabel': self.en2zh.get(self.required, self.required),
                'columns': [{'field': en, 'label': self.en2zh.get(en, en), 'wide': bool(wide)}
                            for en, wide in self.display]}

    # ---- 内部 ----
    def _ensure_id(self, cur):
        cur.execute(f'ALTER TABLE {self.table} ADD COLUMN IF NOT EXISTS id bigserial')
        cur.execute(f"COMMENT ON COLUMN {self.table}.id IS '自增行ID（数据管理在线编辑用，非业务字段）'")

    def _touch(self, cur):
        cur.execute('CREATE TABLE IF NOT EXISTS data_meta (loaded_at timestamptz NOT NULL)')
        cur.execute('DELETE FROM data_meta')
        cur.execute('INSERT INTO data_meta VALUES (now())')

    def _clean(self, fields: dict) -> dict:
        out = {}
        for k, v in fields.items():
            if k not in self.editable:
                continue
            if isinstance(v, str):
                v = v.strip() or None
            out[k] = v
        return out

    def _cell(self, v, typ: str):
        if v is None:
            return None
        if typ == 'text':
            if isinstance(v, float) and v.is_integer():
                return str(int(v))          # 避免数字 ID 变成 "13684045.0"
            s = str(v).strip()
            return s or None
        try:
            return int(float(v)) if typ == 'int' else float(v)
        except (TypeError, ValueError):
            return None

    # ---- CRUD ----
    def list_rows(self) -> list:
        with get_conn() as conn:
            cur = conn.cursor()
            self._ensure_id(cur)
            conn.commit()
            cur.execute(f"""SELECT id, {', '.join(self.en_cols)} FROM {self.table}
                            WHERE {self.not_empty}
                            ORDER BY {self.required} NULLS LAST, id""")
            cols = ['id'] + self.en_cols
            return [dict(zip(cols, r)) for r in cur.fetchall()]

    def insert_row(self, fields: dict) -> int:
        fields = self._clean(fields)
        if not fields.get(self.required):
            raise ValueError(f'{self.en2zh.get(self.required, self.required)}不能为空')
        with get_conn() as conn:
            cur = conn.cursor()
            self._ensure_id(cur)
            cols = list(fields)
            cur.execute(
                f"INSERT INTO {self.table} ({', '.join(cols)}) VALUES ({', '.join(['%s'] * len(cols))}) RETURNING id",
                [fields[c] for c in cols])
            nid = cur.fetchone()[0]
            self._touch(cur)
            conn.commit()
        recon.invalidate()
        return nid

    def update_row(self, rid: int, fields: dict) -> bool:
        fields = self._clean(fields)
        if self.required in fields and not fields[self.required]:
            raise ValueError(f'{self.en2zh.get(self.required, self.required)}不能为空')
        if not fields:
            return False
        with get_conn() as conn:
            cur = conn.cursor()
            self._ensure_id(cur)
            sets = ', '.join(f'{c} = %s' for c in fields)
            cur.execute(f'UPDATE {self.table} SET {sets} WHERE id = %s', [*fields.values(), rid])
            ok = cur.rowcount > 0
            self._touch(cur)
            conn.commit()
        recon.invalidate()
        return ok

    def delete_row(self, rid: int) -> bool:
        with get_conn() as conn:
            cur = conn.cursor()
            self._ensure_id(cur)
            cur.execute(f'DELETE FROM {self.table} WHERE id = %s', (rid,))
            ok = cur.rowcount > 0
            self._touch(cur)
            conn.commit()
        recon.invalidate()
        return ok

    def export_xlsx(self) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        rows = self.list_rows()
        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]
        ws.append([zh for _, zh, _ in self.columns])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='DDEBF7')
        ws.freeze_panes = 'A2'
        for r in rows:
            ws.append([r.get(en) for en in self.en_cols])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_xlsx(self, content: bytes) -> dict:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else '' for h in next(it)]
        except StopIteration:
            raise ValueError('文件为空') from None
        idx = {i: self.zh2en[h] for i, h in enumerate(headers) if h in self.zh2en}
        if self.required not in idx.values():
            raise ValueError(f'缺少"{self.en2zh.get(self.required)}"列，实际表头: {headers[:10]}')
        rows, skipped = [], 0
        for raw in it:
            rec = {en: None for en in self.en_cols}
            for i, en in idx.items():
                if i < len(raw):
                    rec[en] = self._cell(raw[i], self.types[en])
            if not any(rec.get(k) for k in self.key_fields):
                skipped += 1
                continue
            rows.append(rec)
        if not rows:
            raise ValueError('文件里没有有效数据行（关键列全为空）')
        with get_conn() as conn:
            cur = conn.cursor()
            self._ensure_id(cur)
            cur.execute(f'DELETE FROM {self.table}')
            cols = ', '.join(self.en_cols)
            ph = ', '.join(['%s'] * len(self.en_cols))
            cur.executemany(f'INSERT INTO {self.table} ({cols}) VALUES ({ph})',
                            [[r[en] for en in self.en_cols] for r in rows])
            self._touch(cur)
            conn.commit()
        recon.invalidate()
        return {'inserted': len(rows), 'skipped_empty': skipped}


# ======================== 三张表配置 ========================

_MAPPING = TableManager(
    'mapping', 'feishu_store_mapping', '门店映射',
    '飞书《专卖店》主档：门店与三平台 ID 的对应关系。修改立即生效并重算看板；备注填「闭店」或「机场店」的门店不参与核对。',
    columns=[
        ('seq_no', '序号', 'int'), ('province', '省份', 'text'), ('city', '城市', 'text'),
        ('customer_name', '客户名称', 'text'), ('store_name', '门店名称', 'text'),
        ('internal_code', '内部编码', 'text'), ('remark', '备注', 'text'),
        ('jd_name', '京东名称', 'text'), ('jd_id', '京东ID', 'text'),
        ('jd_business_status', '京东营业状态', 'text'), ('meituan_name', '美团名称', 'text'),
        ('meituan_id', '美团ID', 'text'), ('meituan_business_status', '美团营业状态', 'text'),
        ('eleme_name', '饿了么名称', 'text'), ('eleme_id', '饿了么ID', 'text'),
        ('eleme_business_status', '饿了么营业状态', 'text'),
        ('jd_enable_status', '京东启用状态', 'float'), ('verification_status', '验真状态', 'float'),
        ('store_address', '门店地址', 'text'),
    ],
    display=[('customer_name', 1), ('store_name', 1), ('province', 0), ('city', 0),
             ('internal_code', 0), ('remark', 0), ('jd_name', 1), ('jd_id', 0),
             ('meituan_name', 1), ('meituan_id', 0), ('eleme_name', 1), ('eleme_id', 0),
             ('store_address', 0)],
    required='store_name',
    key_fields=('store_name', 'customer_name', 'jd_id', 'meituan_id', 'eleme_id'))

_JD_OUTLET = TableManager(
    'jd_outlet', 'feishu_jd_outlet', '京东网点',
    '京东网点清单（feishu_jd_outlet）。网点保障核对时取「门店状态=启用」的网点，按「经销商」匹配客户、按「门店编号」取京东库存。',
    columns=[
        ('store_code', '门店编号', 'text'), ('store_name', '门店名称', 'text'),
        ('video_status', '视频状态', 'text'), ('store_type', '门店性质', 'text'),
        ('remark', '备注', 'text'), ('dealer', '经销商', 'text'),
        ('merchant_store_code', '商家门店编号', 'text'), ('hourly_store_code', '小时购门店编号', 'text'),
        ('city', '所在城市', 'text'), ('batch', '批次', 'text'), ('district', '行政区', 'text'),
        ('store_address', '门店地址', 'text'), ('business_status', '营业状态', 'text'),
        ('store_status', '门店状态', 'text'), ('hourly_business_status', '小时购营业状态', 'text'),
        ('extra_status', '网点标注', 'text'), ('match_table_status', '匹配表状态', 'text'),
        ('verification_status', '验真状态', 'text'),
    ],
    display=[('store_code', 0), ('store_name', 1), ('dealer', 1), ('store_type', 0),
             ('store_status', 0), ('business_status', 0), ('city', 0), ('district', 0),
             ('store_address', 1), ('batch', 0), ('remark', 0)],
    required='store_name',
    key_fields=('store_name', 'store_code', 'dealer'))

_MT_OUTLET = TableManager(
    'mt_outlet', 'feishu_meituan_outlet', '美团网点',
    '美团网点清单（feishu_meituan_outlet）。网点保障核对时取「营业状态=营业中」的网点，按「经销商」匹配客户、按「门店ID」取美团库存。',
    columns=[
        ('store_name', '门店名称', 'text'), ('store_id', '门店ID', 'text'),
        ('internal_code', '内部编码', 'text'), ('business_status', '营业状态', 'text'),
        ('store_type', '门店类型', 'text'), ('dealer', '经销商', 'text'),
        ('warning_status', '预警情况', 'text'), ('city', '所在城市', 'text'),
        ('contact_phone', '联系电话', 'text'), ('store_address', '门店地址', 'text'),
        ('business_hours', '营业时间', 'text'), ('delivery_method', '配送方式', 'text'),
    ],
    display=[('store_id', 0), ('store_name', 1), ('dealer', 1), ('store_type', 0),
             ('business_status', 0), ('city', 0), ('contact_phone', 0), ('store_address', 1),
             ('business_hours', 1), ('delivery_method', 0)],
    required='store_name',
    key_fields=('store_name', 'store_id', 'dealer'))

REGISTRY = {m.key: m for m in (_MAPPING, _JD_OUTLET, _MT_OUTLET)}
