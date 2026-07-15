# -*- coding: utf-8 -*-
"""门店映射表（feishu_store_mapping）管理：增删改查 + Excel 导入导出。

- 表本身没有主键，这里幂等地补一个自增 id 列（loader 重建表后首次调用会自动补回）
- 导入是"清空重灌"而非删表重建，核对视图不受影响、无需重建
- 任何变更后失效看板数据缓存并刷新 data_meta（看板"数据快照"时间）
"""
import io
from typing import Optional

from ..db import get_conn
from . import recon

TABLE = 'feishu_store_mapping'

# (英文列, 中文表头, 类型) —— 顺序即导出列序，与飞书原表一致
COLUMNS = [
    ('seq_no', '序号', 'int'),
    ('province', '省份', 'text'),
    ('city', '城市', 'text'),
    ('customer_name', '客户名称', 'text'),
    ('store_name', '门店名称', 'text'),
    ('internal_code', '内部编码', 'text'),
    ('remark', '备注', 'text'),
    ('jd_name', '京东名称', 'text'),
    ('jd_id', '京东ID', 'text'),
    ('jd_business_status', '京东营业状态', 'text'),
    ('meituan_name', '美团名称', 'text'),
    ('meituan_id', '美团ID', 'text'),
    ('meituan_business_status', '美团营业状态', 'text'),
    ('eleme_name', '饿了么名称', 'text'),
    ('eleme_id', '饿了么ID', 'text'),
    ('eleme_business_status', '饿了么营业状态', 'text'),
    ('jd_enable_status', '京东启用状态', 'float'),
    ('verification_status', '验真状态', 'float'),
]
EN_COLS = [c for c, _, _ in COLUMNS]
ZH2EN = {zh: en for en, zh, _ in COLUMNS}
TYPES = {en: t for en, _, t in COLUMNS}

# 前端允许编辑的字段（服务端白名单）
EDITABLE = {'province', 'city', 'customer_name', 'store_name', 'internal_code',
            'remark', 'jd_name', 'jd_id', 'jd_business_status',
            'meituan_name', 'meituan_id', 'meituan_business_status',
            'eleme_name', 'eleme_id', 'eleme_business_status'}

# 全空行判定字段（飞书导出里约 1400 行空行）
_KEY_FIELDS = ('store_name', 'customer_name', 'jd_id', 'meituan_id', 'eleme_id')
_NOT_EMPTY = ' OR '.join(f'{c} IS NOT NULL' for c in _KEY_FIELDS)


def _ensure_id(cur):
    cur.execute(f'ALTER TABLE {TABLE} ADD COLUMN IF NOT EXISTS id bigserial')


def _touch(cur):
    """变更后：刷新数据装载时间。调用方 commit 后需 recon.invalidate()。"""
    cur.execute('CREATE TABLE IF NOT EXISTS data_meta (loaded_at timestamptz NOT NULL)')
    cur.execute('DELETE FROM data_meta')
    cur.execute('INSERT INTO data_meta VALUES (now())')


def list_rows() -> list:
    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_id(cur)
        conn.commit()
        cur.execute(f"""SELECT id, {', '.join(EN_COLS)} FROM {TABLE}
                        WHERE {_NOT_EMPTY}
                        ORDER BY customer_name NULLS LAST, store_name NULLS LAST, id""")
        cols = ['id'] + EN_COLS
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def _clean(fields: dict) -> dict:
    out = {}
    for k, v in fields.items():
        if k not in EDITABLE:
            continue
        if isinstance(v, str):
            v = v.strip() or None
        out[k] = v
    return out


def insert_row(fields: dict) -> int:
    fields = _clean(fields)
    if not fields.get('store_name'):
        raise ValueError('门店名称不能为空')
    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_id(cur)
        cols = list(fields)
        cur.execute(
            f"INSERT INTO {TABLE} ({', '.join(cols)}) VALUES ({', '.join(['%s'] * len(cols))}) RETURNING id",
            [fields[c] for c in cols])
        new_id = cur.fetchone()[0]
        _touch(cur)
        conn.commit()
    recon.invalidate()
    return new_id


def update_row(row_id: int, fields: dict) -> bool:
    fields = _clean(fields)
    if 'store_name' in fields and not fields['store_name']:
        raise ValueError('门店名称不能为空')
    if not fields:
        return False
    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_id(cur)
        sets = ', '.join(f'{c} = %s' for c in fields)
        cur.execute(f'UPDATE {TABLE} SET {sets} WHERE id = %s',
                    [*fields.values(), row_id])
        ok = cur.rowcount > 0
        _touch(cur)
        conn.commit()
    recon.invalidate()
    return ok


def delete_row(row_id: int) -> bool:
    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_id(cur)
        cur.execute(f'DELETE FROM {TABLE} WHERE id = %s', (row_id,))
        ok = cur.rowcount > 0
        _touch(cur)
        conn.commit()
    recon.invalidate()
    return ok


def export_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    rows = list_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = '门店映射'
    ws.append([zh for _, zh, _ in COLUMNS])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDEBF7')
    ws.freeze_panes = 'A2'
    for r in rows:
        ws.append([r.get(en) for en in EN_COLS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(v, typ: str):
    if v is None:
        return None
    if typ == 'text':
        if isinstance(v, float) and v.is_integer():
            return str(int(v))          # 避免数字 ID 变成 "13684045.0"
        s = str(v).strip()
        return s or None
    try:
        return int(float(v)) if typ == 'int' else float(v)
    except (TypeError, ValueError):
        return None


def import_xlsx(content: bytes) -> dict:
    """全量替换：清空表后按导出格式的中文表头重灌。不删表，视图不受影响。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else '' for h in next(it)]
    except StopIteration:
        raise ValueError('文件为空') from None
    idx = {}   # 列下标 -> 英文列
    for i, h in enumerate(headers):
        if h in ZH2EN:
            idx[i] = ZH2EN[h]
    if 'store_name' not in idx.values():
        raise ValueError(f'缺少"门店名称"列，实际表头: {headers[:10]}')

    rows, skipped = [], 0
    for raw in it:
        rec = {en: None for en in EN_COLS}
        for i, en in idx.items():
            if i < len(raw):
                rec[en] = _cell(raw[i], TYPES[en])
        if not any(rec.get(k) for k in _KEY_FIELDS):
            skipped += 1
            continue
        rows.append(rec)
    if not rows:
        raise ValueError('文件里没有有效数据行（门店名称等关键列全为空）')

    with get_conn() as conn:
        cur = conn.cursor()
        _ensure_id(cur)
        cur.execute(f'DELETE FROM {TABLE}')
        cols = ', '.join(EN_COLS)
        ph = ', '.join(['%s'] * len(EN_COLS))
        cur.executemany(f'INSERT INTO {TABLE} ({cols}) VALUES ({ph})',
                        [[r[en] for en in EN_COLS] for r in rows])
        _touch(cur)
        conn.commit()
    recon.invalidate()
    return {'inserted': len(rows), 'skipped_empty': skipped}
